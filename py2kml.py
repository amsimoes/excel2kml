from openpyxl import * 
from unidecode import unidecode
import warnings
import sys

colors = {"red":"DB4436", "blue":"4186F0", "green":"009D57",
		  "yellow":"F4EB37"}


# Coordinates "Latitude, Longitude" -> "Longitude, Latitude"
def invert_coords(coords):
	if coords and len(coords) >= 12:
		l = coords.split(",")
		l[0], l[1] = l[1], l[0]
		return str(l[0]+","+l[1])


def file_header(f, out):
	with open(out, "w") as k:
		k.write("<?xml version='1.0' encoding='UTF-8'?>\n<kml xmlns='http://www.opengis.net/kml/2.2'>\n\t<Document>\n\t\t<name>"+f[0]+"</name>\n\t\t<description><![CDATA[]]></description>\n\t\t<Folder>\n\t\t\t<name>"+f[0]+"</name>\n")
		k.close()


def file_ending(out):
	with open("kml_ending.txt") as f:
		with open(out, "a") as k:
			for line in f:
				k.write(line)


def input_data():
	names = input("Column range for points' names (ex: E2:E80): ")
	if len(names) < 5:
		print("Invalid input! Program crashing... brrrzzzztttt boom!!!")
		sys.exit()
	col_coords = input("Coords column letter (Latitude, Longitude): ")
	# Pedir coluna com cores dos pontos
	op = input("Is there a column for points colors? (Yes/No): ")
	if (op == "Yes" or op == "yes" or op == "y"):
		col_colors = input("Colors column letter: ")
		return names, col_coords.upper(), col_colors.upper()
	return names, col_coords.upper(), None


def parser(f):
	wb = load_workbook(filename = f)
	ws = wb.active

	output = input("Output KML filename: ")
	file_header(f.split("."), output+".kml")
	names, col_coords, col_colors = input_data()

	# default color -> Red
	if not col_colors:
		color = "DB4436"

	for row in ws.iter_rows(names):
		for cell in row:
			title = unidecode(cell.value)
			coords = ws[col_coords+str(cell.row)].value
			if col_colors:
				color = colors[ws[col_colors+str(cell.row)].value.lower()]
			inv_coords = invert_coords(coords)
			with open(output+".kml", "a") as k:
				k.write("\t\t\t<Placemark>\n\t\t\t\t<name>"+title+"</name>\n\t\t\t\t<styleUrl>#icon-503-"+color+"-nodesc</styleUrl>\n\t\t\t\t<Point>\n\t\t\t\t\t<coordinates>"+inv_coords+",0.0</coordinates>\n\t\t\t\t</Point>\n\t\t\t</Placemark>\n")
				k.close()
	file_ending(output+".kml")
	print("\n*** Congratulations! Your file "+output+".kml was generated successfully! ***")


if __name__ == '__main__':
	warnings.filterwarnings("ignore")
	path = input("XLS file: ")
	parser(path+".xlsx")